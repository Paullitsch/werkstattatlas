#!/usr/bin/env python3
"""
Werkstattatlas.info Scraper
Scrapt alle Werkstätten aus der Kategorie "Wartung, Instandhaltung, Service"
und speichert sie als JSON.
"""

import json
import re
import sys
import time
from pathlib import Path
from urllib.request import urlopen, Request
from html.parser import HTMLParser

BASE_URL = "http://www.werkstattatlas.info"
LIST_URL = f"{BASE_URL}/unternehmen/283-wartung_instandhaltung_service.html"
OUTPUT_JSON = Path("data/werkstaetten.json")
OUTPUT_XLSX = Path("data/werkstaetten.xlsx")
DELAY = 1.0  # Seconds between requests (be polite)


class ListPageParser(HTMLParser):
    """Parse listing page: extract entries (name, url, address) and page count."""

    def __init__(self):
        super().__init__()
        self.entries: list[dict] = []
        self.max_page = 1
        self._in_h2_lead = False
        self._in_entry_p = False
        self._current: dict = {}
        self._after_h2 = False

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        # h2.lead > a = entry link
        if tag == "h2" and "lead" in a.get("class", ""):
            self._in_h2_lead = True
            self._current = {}
        elif tag == "a" and self._in_h2_lead:
            href = a.get("href", "")
            if href.startswith("/unternehmen/") and href != "/unternehmen.html":
                self._current["url"] = BASE_URL + href
                # Extract ID from URL
                m = re.search(r"/(\d+)-", href)
                if m:
                    self._current["id"] = int(m.group(1))
        # p right after h2 = address
        elif tag == "p" and self._after_h2:
            self._in_entry_p = True
        # Pagination links
        elif tag == "a":
            href = a.get("href", "")
            m = re.search(r"site=(\d+)", href)
            if m:
                self.max_page = max(self.max_page, int(m.group(1)))

    def handle_endtag(self, tag):
        if tag == "h2" and self._in_h2_lead:
            self._in_h2_lead = False
            self._after_h2 = True
        elif tag == "p" and self._in_entry_p:
            self._in_entry_p = False
            self._after_h2 = False
            if self._current.get("name"):
                self.entries.append(self._current)
                self._current = {}

    def handle_data(self, data):
        text = data.strip()
        if not text:
            return
        if self._in_h2_lead:
            self._current["name"] = text
        elif self._in_entry_p:
            self._current["address_raw"] = text
            self._parse_address(text)

    def _parse_address(self, raw: str):
        """Parse address in various formats from werkstattatlas."""
        parts = [p.strip() for p in raw.split("|")]

        # Try to find location part in any segment
        # Formats: "DE-10317 Berlin (D)", "3088 GC Rotterdam (NL)", "PL-85-082 Bydgoszcz"
        # "GB-DN4 5PN Doncaster", "SE-724 65 Västeras", "CZ-751 52 Prerov"
        location_idx = -1
        for i, part in enumerate(parts):
            if re.search(r"[A-Z]{2}-?\d{3,5}", part) or re.search(r"\d{4,5}\s+[A-Z]{2}\s+\w", part):
                location_idx = i
                break

        if location_idx >= 0:
            location = parts[location_idx]
            # Try multiple formats
            m = re.match(r"([A-Z]{2})-?(\d{4,5})\s+(.+?)(?:\s*\((.+?)\))?$", location)
            if not m:
                # "3088 GC Rotterdam (Niederlande)" — NL format
                m2 = re.match(r"(\d{4})\s+([A-Z]{2})\s+(.+?)(?:\s*\((.+?)\))?$", location)
                if m2:
                    m = None  # handle below
                    self._current["country_code"] = "NL"
                    self._current["postal_code"] = m2.group(1) + " " + m2.group(2)
                    self._current["city"] = m2.group(3).strip()
                    self._current["country"] = m2.group(4) or ""
            if not m and "country_code" not in self._current:
                # "GB-DN4 5PN Doncaster" or "CZ-751 52 Prerov" or "SE-724 65 Västeras"
                m3 = re.match(r"([A-Z]{2})-?([\w\d][\w\d -]{2,8}?)\s+([A-ZÄÖÜ].+?)(?:\s*\((.+?)\))?$", location)
                if m3:
                    m = None
                    self._current["country_code"] = m3.group(1)
                    self._current["postal_code"] = m3.group(2).strip()
                    self._current["city"] = m3.group(3).strip()
                    self._current["country"] = m3.group(4) or ""
            if m:
                self._current["country_code"] = m.group(1)
                self._current["postal_code"] = m.group(2)
                self._current["city"] = m.group(3).strip()
                self._current["country"] = m.group(4) or ""

            # Parts before location = workshop name + street
            before = parts[:location_idx]
            if len(before) >= 2:
                self._current["workshop_name"] = before[0]
                self._current["street"] = before[1]
            elif len(before) == 1:
                # Could be workshop name or street — check if it looks like a street
                if re.search(r"\d", before[0]) and re.search(r"straße|weg|gasse|platz|ring|allee", before[0], re.I):
                    self._current["street"] = before[0]
                else:
                    self._current["workshop_name"] = before[0]

            # Parts after location (rare)
            after = parts[location_idx + 1:]
            if after and not self._current.get("street"):
                self._current["street"] = after[0]
        else:
            # No location found — just store parts
            if len(parts) >= 2:
                self._current["workshop_name"] = parts[0]
                self._current["street"] = parts[1]
            elif len(parts) == 1:
                self._current["workshop_name"] = parts[0]


class DetailPageParser(HTMLParser):
    """Parse detail page: extract phone, fax, website."""

    def __init__(self):
        super().__init__()
        self.phone = ""
        self.fax = ""
        self.website = ""
        self._capture_text = False
        self._buffer = ""

    def handle_data(self, data):
        self._buffer += data

    def close(self):
        super().close()
        # Extract from buffer
        for line in self._buffer.split("\n"):
            line = line.strip()
            if line.startswith("Telefon:"):
                self.phone = line.replace("Telefon:", "").strip()
            elif line.startswith("Telefax:"):
                self.fax = line.replace("Telefax:", "").strip()

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "a":
            href = a.get("href", "")
            text = a.get("title", "")
            # Website link (not internal, not google, not search)
            if href.startswith("http") and "werkstattatlas" not in href and "google" not in href and "sobipro" not in href:
                if not self.website:
                    self.website = href


def fetch(url: str) -> str:
    """Fetch URL with User-Agent header."""
    req = Request(url, headers={"User-Agent": "WerkstattScraper/1.0 (github.com/Paullitsch/werkstattatlas)"})
    with urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8", errors="replace")


def scrape_list_page(page: int) -> tuple[list[dict], int]:
    """Scrape one listing page. Returns (entries, max_page)."""
    url = f"{LIST_URL}?site={page}"
    html = fetch(url)
    parser = ListPageParser()
    parser.feed(html)
    return parser.entries, parser.max_page


def scrape_detail(url: str) -> dict:
    """Scrape detail page for phone/fax/website."""
    html = fetch(url)
    parser = DetailPageParser()
    parser.feed(html)
    parser.close()
    result = {}
    if parser.phone:
        result["phone"] = parser.phone
    if parser.fax:
        result["fax"] = parser.fax
    if parser.website:
        result["website"] = parser.website
    return result


XLSX_COLUMNS = [
    ("id", "ID"),
    ("name", "Unternehmen"),
    ("workshop_name", "Werkstatt"),
    ("street", "Straße"),
    ("postal_code", "PLZ"),
    ("city", "Stadt"),
    ("country_code", "Land"),
    ("country", "Land (Name)"),
    ("phone", "Telefon"),
    ("fax", "Fax"),
    ("website", "Website"),
    ("url", "Werkstattatlas-URL"),
]


def export_xlsx(entries: list[dict]) -> None:
    """Export entries to Excel file with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Werkstätten"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Write headers
    for col, (key, label) in enumerate(XLSX_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data
    for row, entry in enumerate(entries, 2):
        for col, (key, _) in enumerate(XLSX_COLUMNS, 1):
            val = entry.get(key, "")
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if key == "url" or key == "website":
                cell.alignment = Alignment(wrap_text=False)

    # Auto-width columns
    for col, (key, label) in enumerate(XLSX_COLUMNS, 1):
        max_len = len(label)
        for row in range(2, min(len(entries) + 2, 50)):  # Sample first 50
            val = str(ws.cell(row=row, column=col).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 45)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(XLSX_COLUMNS)).column_letter}{len(entries) + 1}"

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def main():
    skip_details = "--no-details" in sys.argv
    
    print("🔍 Scraping werkstattatlas.info ...")
    
    # Phase 1: Scrape all list pages
    all_entries: list[dict] = []
    seen_ids: set[int] = set()
    
    # First page to get total pages
    entries, max_page = scrape_list_page(1)
    for e in entries:
        eid = e.get("id", 0)
        if eid not in seen_ids:
            seen_ids.add(eid)
            all_entries.append(e)
    
    print(f"  Seite 1/{max_page}: {len(entries)} Einträge")
    
    for page in range(2, max_page + 1):
        time.sleep(DELAY)
        entries, _ = scrape_list_page(page)
        new = 0
        for e in entries:
            eid = e.get("id", 0)
            if eid not in seen_ids:
                seen_ids.add(eid)
                all_entries.append(e)
                new += 1
        print(f"  Seite {page}/{max_page}: {len(entries)} Einträge ({new} neu)")
    
    print(f"\n📋 {len(all_entries)} Werkstätten gefunden")
    
    # Phase 2: Scrape detail pages (optional)
    if not skip_details:
        print("\n🔎 Scrape Detail-Seiten ...")
        for i, entry in enumerate(all_entries):
            url = entry.get("url", "")
            if not url:
                continue
            try:
                time.sleep(DELAY)
                details = scrape_detail(url)
                entry.update(details)
                if (i + 1) % 25 == 0:
                    print(f"  {i + 1}/{len(all_entries)} ...")
            except Exception as e:
                print(f"  ⚠️ Fehler bei {url}: {e}")
    
    # Sort by ID
    all_entries.sort(key=lambda e: e.get("id", 0))
    
    # Save JSON
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(all_entries, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ JSON: {OUTPUT_JSON} ({len(all_entries)} Einträge)")
    
    # Save XLSX
    export_xlsx(all_entries)
    print(f"✅ XLSX: {OUTPUT_XLSX}")
    
    # Stats
    countries: dict[str, int] = {}
    for e in all_entries:
        cc = e.get("country_code", "??")
        countries[cc] = countries.get(cc, 0) + 1
    print("\n📊 Länder:")
    for cc, count in sorted(countries.items(), key=lambda x: -x[1]):
        print(f"  {cc}: {count}")


if __name__ == "__main__":
    main()
